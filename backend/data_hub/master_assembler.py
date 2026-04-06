"""Master File Assembler — creates a temporary xlsx workbook from uploaded source files
that the original dashboard_refresh_all_in_one.py processor can read.

This bridges the gap between individual source file uploads and the
original 3400-line processor that expects a Master File workbook."""

import os
import tempfile
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime


def assemble_master_xlsx(cache_dir, template_path=None):
    """
    Assemble uploaded source DataFrames into a temporary xlsx file
    with the sheet names the dashboard processor expects.

    All pre-computed sheets (RS, DPD, INV, OBJ, HIST, LK, etc.) are
    populated from the source data, replicating the Master File's pivot tables.

    Args:
        cache_dir: Path to cache directory with data/*.parquet files
        template_path: Optional path to Dashboard HTML template (for OBJ extraction)

    Returns: path to temporary xlsx file
    """
    from data_hub.sheet_builders import (
        _parse_export_rows, build_retail_sales_sheet, build_dpd_sheet,
        build_inventory_sheet, build_objectives_sheet, build_historical_sheet,
        build_lead_kpis_sheet, build_santander_sheets, build_ga4_sheet_formatted,
    )

    def load(name):
        path = os.path.join(cache_dir, 'data', f'{name}.parquet')
        if os.path.exists(path):
            return pd.read_parquet(path)
        return None

    def load_any(*names):
        """Try loading from multiple possible key names."""
        for name in names:
            df = load(name)
            if df is not None:
                print(f"  Loaded {name}: {len(df)} rows")
                return df
        return None

    sap = load_any('sap_export')
    if sap is None:
        raise RuntimeError("SAP Vehicle Export not uploaded yet.")

    handover = load_any('sap_handover', 'handover')
    leads = load_any('leads', 'c4c_leads')
    stock_pipeline = load_any('stock_pipeline')
    sales_order = load_any('sales_order')
    urban_science = load_any('urban_science')
    campaign_codes = load_any('campaign_codes')
    incentive_spend = load_any('incentive_spend')
    qm_leads = load_any('qm_leads')

    # Auto-detect template path if not provided
    if template_path is None:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        for candidate in [
            os.path.join(base, 'templates', 'dashboard_template.html'),
            os.path.join(base, 'outputs', 'Americas_Daily_Dashboard.html'),
        ]:
            if os.path.exists(candidate):
                template_path = candidate
                break

    # Pre-parse export rows for sheet builders
    print("  Parsing export rows for sheet builders...")
    export_rows, mkt_map = _parse_export_rows(sap, handover, sales_order, campaign_codes, template_path)
    print(f"  Parsed {len(export_rows)} export rows, {len(mkt_map)} dealers mapped")

    # Create temporary xlsx
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()

    wb = openpyxl.Workbook()

    # ═══ SHEET: Export ═══
    ws = wb.active
    ws.title = "Export"
    _write_export_sheet(ws, sap, handover, stock_pipeline, sales_order,
                        campaign_codes, incentive_spend, urban_science)
    print(f"  Export: {len(sap)} rows")

    # ═══ SHEET: Raw Lead Data ═══
    if leads is not None and len(leads) > 0:
        ws_leads = wb.create_sheet("Raw Lead Data")
        _write_leads_sheet(ws_leads, leads)
        print(f"  Raw Lead Data: {len(leads)} rows")

    # ═══ SHEET: RBM Assignments ═══
    ws_rbm = wb.create_sheet("RBM Assignments")
    _write_rbm_sheet(ws_rbm, sap, mkt_map)

    # ═══ SHEET: Dealer Address ═══
    ws_addr = wb.create_sheet("Dealer Address")
    ws_addr.append(["Dealer", "Street", "City", "State", "Zip", "Lat", "Lon"])

    # ═══ SHEET: Retail Sales Report ═══
    ws_rsr = wb.create_sheet("Retail Sales Report")
    build_retail_sales_sheet(ws_rsr, export_rows, mkt_map)
    print("  Retail Sales Report: populated from export data")

    # ═══ SHEET: Dealer Performance Dashboard ═══
    ws_dpd = wb.create_sheet("Dealer Performance Dashboard")
    build_dpd_sheet(ws_dpd, export_rows, mkt_map, leads)
    print("  Dealer Performance Dashboard: populated from export data")

    # ═══ SHEET: Dealer Inventory Report ═══
    ws_inv = wb.create_sheet("Dealer Inventory Report")
    build_inventory_sheet(ws_inv, export_rows, mkt_map)
    print("  Dealer Inventory Report: populated from export data")

    # ═══ SHEET: Objectives ═══
    ws_obj = wb.create_sheet("Objectives")
    obj_data = build_objectives_sheet(ws_obj, template_path)
    if obj_data:
        print(f"  Objectives: extracted from template ({len(obj_data)} categories)")
    else:
        print("  Objectives: no template data available")

    # ═══ SHEET: Historical Sales ═══
    ws_hist = wb.create_sheet("Historical Sales")
    build_historical_sheet(ws_hist, export_rows, mkt_map)
    print("  Historical Sales: populated from export data")

    # ═══ SHEET: Lead Handling KPIs ═══
    ws_lk = wb.create_sheet("Lead Handling KPIs")
    build_lead_kpis_sheet(ws_lk, leads, mkt_map)
    print(f"  Lead Handling KPIs: {'populated' if leads is not None else 'empty'}")

    # ═══ SHEET: Matchback Report ═══
    ws_mb = wb.create_sheet("Matchback Report")
    for i in range(5):
        ws_mb.append([""] * 20)

    # ═══ GA4 SHEETS ═══
    for ga4_name, sheet_name in [
        ('ga4_engagement', 'G - Engagement Overview'),
        ('ga4_acquisition', 'G - Acquisition Overview'),
        ('ga4_user_attributes', 'G - User Attributes'),
        ('ga4_demographics', 'G - Demographics'),
        ('ga4_tech', 'G - Tech'),
        ('ga4_audiences', 'G - Audiences'),
    ]:
        ga4_df = load(ga4_name)
        ws_ga4 = wb.create_sheet(sheet_name)
        if ga4_df is not None and len(ga4_df) > 0:
            build_ga4_sheet_formatted(ws_ga4, ga4_df, ga4_name)
            print(f"  {sheet_name}: {len(ga4_df)} rows")
        else:
            for i in range(10):
                ws_ga4.append([""] * 10)

    # ═══ SANTANDER SHEETS ═══
    for name in ["Santander Report ", "Santander Report Finance", "Santander Report Lease"]:
        ws_san = wb.create_sheet(name)
        for i in range(5):
            ws_san.append([""] * 30)
    build_santander_sheets(wb, cache_dir)
    print("  Santander: populated from cached data")

    wb.save(tmp.name)
    wb.close()

    print(f"  Assembled Master xlsx: {os.path.getsize(tmp.name) / 1024:.0f} KB")
    return tmp.name


def _write_export_sheet(ws, sap, handover, stock_pipeline, sales_order,
                        campaign_codes, incentive_spend, urban_science):
    """Write the Export sheet in the format the processor expects.

    The processor reads columns by index — ALL indices must match exactly.
    See column map below. The processor iterates export_rows starting at
    row index 2 (rows 0-1 are headers).

    Column map (processor index → field):
     [0]  Customer Full Name          [1]  (Ship to Party)
     [3]  SO Sales Order No           [6]  Invoice Date (fallback HO date)
     [7]  Material Desc               [8]  Vehicle VIN
     [11] Country Name                [12] Status Code
     [13] Status Text                 [14] Channel
     [18] MSRP                        [19] Trim
     [20] Rough Pack / Package        [21] Ext Color
     [22] Seats                       [23] Roof
     [24] Safari Windows              [25] Wheels
     [26] Tyres                       [27] Frame Color
     [28-49] Option columns (22 options: tow_ball, seat_heating, diff_locks_rf,
             access_ladder, aux_battery, aux_switchbar, carpet_mats, compass,
             diff_lock_central, safety_package, floor_trim, front_winch,
             utility_rails, privacy_glass, raised_air_intake, smokers_pack,
             spare_wheel_container, tow_plate_front, rubber_bump_strips,
             steering_wheel, wheel_locks, speaker_system)
     [50] Plant Code                  [51] Handover Date (serial date)
     [52] ETA                         [53] Vessel
     [54] Market Override             [55] Rev Rec Date
     [57] DIS (Days in Stock)         [58] Bill To Dealer
     [72] Variable Spend              [75] Campaign Code
    """
    # Header rows (processor skips rows 0-1)
    ws.append(["Header Row 0"] + [""] * 79)
    ws.append(["Header Row 1"] + [""] * 79)

    # ── Build lookup maps from supplementary data sources ──

    # VIN → handover data
    ho_map = {}
    if handover is not None and len(handover) > 0:
        for _, r in handover.iterrows():
            vin = str(r.get('vin', '')).strip().upper()
            if len(vin) > 3:
                ho_map[vin] = r

    # Order No → vessel/ETA (Stock & Pipeline)
    vessel_map = {}
    if stock_pipeline is not None and len(stock_pipeline) > 0:
        for _, r in stock_pipeline.iterrows():
            on = str(r.get('order_no', '')).strip()
            if len(on) > 3:
                vessel_map[on] = r

    # VIN → bill-to-dealer (Sales Order / List of Sales Orders)
    billto_map = {}
    if sales_order is not None and len(sales_order) > 0:
        for _, r in sales_order.iterrows():
            vin = str(r.get('vin', '')).strip().upper()
            if len(vin) > 3:
                billto_map[vin] = str(r.get('bill_to_dealer', r.get('customer_name', ''))).strip()

    # VIN → campaign type (CVP / Demo)
    cvp_vins = set()
    demo_vins = set()
    campaign_code_map = {}  # VIN → campaign code string
    if campaign_codes is not None and len(campaign_codes) > 0:
        if 'campaign_type' in campaign_codes.columns:
            cvp_vins = set(campaign_codes[campaign_codes['campaign_type'] == 'CVP']['vin'].dropna().astype(str).str.upper())
            demo_vins = set(campaign_codes[campaign_codes['campaign_type'] == 'Demo']['vin'].dropna().astype(str).str.upper())
        if 'vin' in campaign_codes.columns:
            for _, cr in campaign_codes.iterrows():
                cv = str(cr.get('vin', '')).strip().upper()
                cc = str(cr.get('campaign_code', cr.get('code', ''))).strip()
                if cv and cc:
                    campaign_code_map[cv] = cc

    # VIN → incentive/variable spend
    spend_map = {}
    if incentive_spend is not None and len(incentive_spend) > 0:
        if 'vin' in incentive_spend.columns:
            for _, sr in incentive_spend.iterrows():
                sv = str(sr.get('vin', '')).strip().upper()
                amt = sr.get('spend_amount', sr.get('amount', 0))
                if sv:
                    spend_map[sv] = amt

    # 22 option column names in order (maps to cols 28-49)
    OPTION_COLS = [
        'tow_ball', 'seat_heating', 'diff_locks_rf', 'access_ladder',
        'aux_battery', 'aux_switchbar', 'carpet_mats', 'compass',
        'diff_lock_central', 'safety_package', 'floor_trim', 'front_winch',
        'utility_rails', 'privacy_glass', 'raised_air_intake', 'smokers_pack',
        'spare_wheel_container', 'tow_plate_front', 'rubber_bump_strips',
        'steering_wheel', 'wheel_locks', 'speaker_system',
    ]

    def _sap_val(row, *keys):
        """Get first non-empty value from SAP row by internal or original column name."""
        for k in keys:
            v = row.get(k, None)
            if v is not None and str(v).strip() not in ('', 'nan', 'None'):
                return str(v).strip()
        return ''

    def _date_to_serial(d):
        """Convert a date/datetime/Timestamp to Excel serial number for the processor.

        The processor's serial_to_date() expects: int(float(value)) → days since 1899-12-30.
        Returns None for null/NaT/invalid dates.
        """
        import pandas as _pd
        if d is None:
            return None
        # Handle pandas NaT (Not a Time) — must check before float/hasattr
        if isinstance(d, type(_pd.NaT)) or _pd.isna(d) if not isinstance(d, str) else False:
            return None
        if isinstance(d, (int, float)):
            if np.isnan(d):
                return None
            return d  # Already a serial number
        try:
            from datetime import datetime as dt
            if isinstance(d, str):
                # Try parsing as pandas timestamp first (handles many formats)
                try:
                    d = _pd.to_datetime(d, errors='coerce')
                    if _pd.isna(d):
                        return None
                except Exception:
                    return None
            if hasattr(d, 'toordinal'):
                # Excel serial date: days since 1899-12-30
                delta = d - dt(1899, 12, 30)
                return delta.days + (getattr(delta, 'seconds', 0) / 86400.0)
        except Exception:
            pass
        return None

    for _, r in sap.iterrows():
        vin = _sap_val(r, 'vin', 'Vehicle VIN').upper()
        order_no = _sap_val(r, 'order_no', 'SO Sales Order No')
        customer = _sap_val(r, 'customer_name', 'Customer Full Name')
        material = _sap_val(r, 'material', 'Material Desc')
        country = _sap_val(r, 'country', 'Country Name')
        status_code = _sap_val(r, 'status_code', 'Vehicle Current Primary Status Code')
        status = _sap_val(r, 'status', 'Vehicle Current Primary Status Text (groups)')
        channel = _sap_val(r, 'channel', 'SO Channel Desc')
        msrp = r.get('msrp', r.get('MSRP (US$)', 0))
        trim = _sap_val(r, 'trim', 'Trim Levels Groups')
        rough_pack = _sap_val(r, 'rough_pack', 'Rough Pack Desc')
        ext_color = _sap_val(r, 'ext_color', 'Exterior Paint Colour Desc')
        seats = _sap_val(r, 'seats', 'Seats Material Desc')
        roof = _sap_val(r, 'roof_color', 'Exterior Contrast Roof Colour Desc')
        safari = _sap_val(r, 'safari_windows', 'Safari Windows Desc')
        wheels = _sap_val(r, 'wheels', 'Wheels Desc')
        tyres = _sap_val(r, 'tyres', 'Std Tyre  Opt Tyre Desc')
        frame_color = _sap_val(r, 'frame_color', 'Exterior Ladder Frame Colour Desc')
        plant = _sap_val(r, 'plant_code', 'Plant Code')
        ship_to = _sap_val(r, 'ship_to_party', 'Ship to Party No')
        invoice_date = r.get('invoice_date', r.get('SO BD Date (Invoice Date)', None))

        # Handover data
        ho = ho_map.get(vin, {})
        ho_date = ho.get('handover_date', None) if isinstance(ho, dict) else getattr(ho, 'handover_date', None)
        rev_rec = ho.get('rev_rec_date', None) if isinstance(ho, dict) else getattr(ho, 'rev_rec_date', None)

        # Vessel data (keyed by order_no)
        vsl = vessel_map.get(order_no, {})
        eta = vsl.get('shipping_eta', None) if isinstance(vsl, dict) else getattr(vsl, 'shipping_eta', None)
        vessel = vsl.get('vessel', '') if isinstance(vsl, dict) else getattr(vsl, 'vessel', '')
        dis = vsl.get('days_in_stock', vsl.get('dis', '')) if isinstance(vsl, dict) else getattr(vsl, 'days_in_stock', getattr(vsl, 'dis', ''))

        # Bill-to dealer — from Sales Order, with channel-based fallback
        # The processor classifies "Fleet", "Internal", "Enterprise" as non-retail
        bill_to = billto_map.get(vin, '')
        if not bill_to:
            # Fallback: derive from SAP channel if no sales order data
            ch = channel.upper()
            if any(x in ch for x in ('FLEET', 'RENTAL')):
                bill_to = 'Fleet'
            elif any(x in ch for x in ('INTERNAL', 'EMPLOYEE')):
                bill_to = 'Internal'
            elif 'ENTERPRISE' in ch or 'IECP' in ch:
                bill_to = 'Enterprise'
            else:
                bill_to = 'Not Handed Over'

        # Market area (from SAP data if available)
        market = _sap_val(r, 'market_area', 'region_group', 'Country Region Group')

        # CVP/Demo flags
        cvp = 'Yes' if vin in cvp_vins else 'No'
        demo = 'Yes' if vin in demo_vins else 'No'

        # Incentive spend
        var_spend = spend_map.get(vin, '')

        # Campaign code
        camp_code = campaign_code_map.get(vin, '')

        # Convert dates to Excel serial numbers for the processor's serial_to_date()
        ho_serial = _date_to_serial(ho_date)
        rr_serial = _date_to_serial(rev_rec)
        eta_serial = _date_to_serial(eta)
        inv_serial = _date_to_serial(invoice_date)

        # Build row — 80 columns to match Export sheet layout
        # ALL indices must match what the processor reads
        row = [''] * 80
        row[0] = customer          # [0]  Customer Full Name
        row[1] = ship_to           # [1]  Ship to Party No
        row[3] = order_no          # [3]  SO Sales Order No
        row[6] = inv_serial        # [6]  Invoice Date (serial, fallback HO date)
        row[7] = material          # [7]  Material Desc (body type + MY detection)
        row[8] = vin               # [8]  Vehicle VIN
        row[11] = country          # [11] Country Name
        row[12] = status_code      # [12] Status Code
        row[13] = status           # [13] Status Text (groups)
        row[14] = channel          # [14] Channel
        row[18] = msrp             # [18] MSRP
        row[19] = trim             # [19] Trim
        row[20] = rough_pack       # [20] Rough Pack / Package
        row[21] = ext_color        # [21] Exterior Color
        row[22] = seats            # [22] Seats Material
        row[23] = roof             # [23] Roof Color
        row[24] = safari           # [24] Safari Windows
        row[25] = wheels           # [25] Wheels
        row[26] = tyres            # [26] Tyres
        row[27] = frame_color      # [27] Frame Color

        # Options (cols 28-49)
        for i, opt_col in enumerate(OPTION_COLS):
            row[28 + i] = _sap_val(r, opt_col)

        row[50] = plant            # [50] Plant Code
        row[51] = ho_serial        # [51] Handover Date (serial number)
        row[52] = eta_serial       # [52] ETA (serial number)
        row[53] = vessel           # [53] Vessel
        row[54] = market           # [54] Market Override
        row[55] = rr_serial        # [55] Rev Rec Date (serial number)
        row[57] = dis              # [57] Days in Stock / DIS
        row[58] = bill_to          # [58] Bill To Dealer
        row[62] = cvp              # [62] CVP flag
        row[63] = demo             # [63] Demo flag
        row[72] = var_spend        # [72] Variable/Incentive Spend
        row[75] = camp_code        # [75] Campaign Code

        ws.append(row)


def _write_leads_sheet(ws, leads):
    """Write Raw Lead Data sheet in ORIGINAL C4C column order.

    The processor reads Raw Lead Data by column INDEX (not name).
    Critical indices:
      [2]  = Retailer Name (dealer)
      [12] = Marketing Unit (market)
      [16] = Start Date (lead creation date — serial number!)
      [25] = Test drive booking date (serial number!)
      [39] = QM flag (added by us, may not exist in original)

    The columns must be in the same order as the original C4C export.
    Dates must be converted to Excel serial numbers for serial_to_date().
    """
    # Original C4C column order → internal DataFrame column names
    LEADS_COL_ORDER = [
        ('Lead ID', 'lead_id'),                          # [0]
        ('Name', 'lead_name'),                           # [1]
        ('Retailer Name', 'retailer_name'),              # [2] ← processor uses
        ('Company/Customer', 'customer_name'),           # [3]
        ('Customer Phone', 'customer_phone'),            # [4]
        ('Customer Mobile', 'customer_mobile'),          # [5]
        ('Customer E-Mail', 'customer_email'),           # [6]
        ('Status', 'lead_status'),                       # [7]
        ('Reason Code', 'reason_code'),                  # [8]
        ('Retailer Status', 'retailer_status'),          # [9]
        ('Retailer Country Name', 'retailer_country'),   # [10]
        ('Country/Region', 'country_region'),            # [11]
        ('Marketing Unit', 'marketing_unit'),            # [12]
        ('Source', 'source'),                             # [13]
        ('Qualified', 'qualified_date'),                  # [14]
        ('Closed', 'closed_date'),                       # [15]
        ('Start Date', 'start_date'),                    # [16] ← processor uses as lead creation date
        ('End Date', 'end_date'),                        # [17]
        ('Category', 'category'),                        # [18]
        ('Owner', 'owner'),                              # [19]
        ('Created On', 'created_on'),                    # [20]
        ('Model of Interest', 'model_interest'),         # [21]
        ('Test Drive Requested Date', 'td_requested'),   # [22]
        ('First contact attempt', 'first_contact'),      # [23]
        ('Retailer First Status Changed On', 'first_status_change'),  # [24]
        ('Test drive booking date', 'td_booking_date'),  # [25] ← processor uses
        ('Test drive booking time', 'td_booking_time'),  # [26]
        ('Booking ID', 'booking_id'),                    # [27]
        ('Test Drive Completed Date', 'td_completed_date'),  # [28]
        ('Test drive completed', 'td_completed_flag'),   # [29]
        ('Note Exists', 'note_exists'),                  # [30]
        ('Marketing Unit 2', 'marketing_unit'),          # [31] ← processor reads for market fallback
    ]

    # Date columns that need serial number conversion (for serial_to_date())
    DATE_COLS = {
        'qualified_date', 'closed_date', 'start_date', 'end_date',
        'created_on', 'td_requested', 'first_contact', 'first_status_change',
        'td_booking_date', 'td_completed_date',
    }

    def _to_serial(val):
        """Convert date to Excel serial number."""
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return None
        if isinstance(val, (int, float)):
            return val  # Already serial
        try:
            if isinstance(val, str):
                import pandas as _pd
                val = _pd.to_datetime(val, errors='coerce')
                if _pd.isna(val):
                    return None
            if hasattr(val, 'toordinal'):
                delta = val - datetime(1899, 12, 30)
                return delta.days + (getattr(delta, 'seconds', 0) / 86400.0)
        except Exception:
            pass
        return None

    # Write header row (original names)
    headers = [orig for orig, _ in LEADS_COL_ORDER]
    # Pad to 40 columns to ensure r[39] access works
    while len(headers) < 40:
        headers.append('')
    ws.append(headers)

    # Write data rows in the correct column order
    for _, r in leads.iterrows():
        row = []
        for _, int_name in LEADS_COL_ORDER:
            val = r.get(int_name, '')
            if int_name in DATE_COLS:
                serial = _to_serial(val)
                row.append(serial if serial is not None else '')
            else:
                row.append(str(val) if val is not None and str(val) not in ('nan', 'NaT', 'None') else '')

        # Pad to 40 columns so r[39] (QM flag) is accessible
        while len(row) < 40:
            row.append('')

        # Set QM flag at index 39 if available
        body_type = str(r.get('body_type', '')).lower()
        row[39] = 'Yes' if body_type == 'qm' else ''

        ws.append(row)


def _write_rbm_sheet(ws, sap, mkt_map):
    """Write RBM Assignments using the dealer→market mapping.

    The processor reads from row 5+, dealer=col3, market=col5.
    We use mkt_map (extracted from template) instead of SAP's market_area
    which is just "AMERICAS" for all dealers.
    """
    for i in range(5):
        ws.append([""] * 10)

    # Write dealer→market assignments from our map
    written = set()
    if 'customer_name' in sap.columns:
        for _, r in sap.iterrows():
            raw_name = str(r['customer_name']).strip()
            norm = raw_name.replace(' INEOS Grenadier', '').replace(' INEOS', '').replace(' GRENADIER', '').strip()
            norm = ' '.join(w for w in norm.split() if w.upper() != 'GRENADIER').strip()
            upper = norm.upper()

            if upper in written:
                continue

            market = mkt_map.get(upper, '')
            if not market:
                # Fuzzy match
                for k, v in mkt_map.items():
                    if upper in k or k in upper:
                        market = v
                        break
            if market:
                row = [""] * 10
                row[3] = norm
                row[5] = market
                ws.append(row)
                written.add(upper)


def _write_ga4_sheet(ws, df, ga4_type):
    """Write GA4 data in the format the processor expects."""
    # GA4 sheets have metadata rows then data
    # The processor reads specific columns based on the sheet
    for i in range(9):
        ws.append([""] * 10)

    # Write data rows
    cols = list(df.columns)
    for _, r in df.iterrows():
        row = [r.get(c, '') for c in cols]
        ws.append(row)
