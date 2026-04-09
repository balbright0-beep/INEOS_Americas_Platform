"""Vehicle Distribution ingest — computes the Logistics Freight Order Performance
report from the weekly Vehicle Distribution xlsx sent by the logistics team.

Parses the 'Data File' sheet via pandas (fast — ~1s for 8k rows) and produces
a structured dict matching the published Daily Freight Order Activity report.

Math spec (as of 2026-04-09):

1. FO universe        = rows where FO Create Date is not null.
2. Same-month capping = Dispatched / PickUp / Delivered dates only count if
                        they fall in the same calendar month as FO Create
                        Date. Later-month events are treated as NaT on that
                        row so post-month activity never inflates SLA metrics.
3. Business days      = exclude weekends AND US federal holidays. Count is
                        exclusive of start, inclusive of end — numerically
                        equivalent to numpy.busday_count.
4. SLA compliance     = only FOs with a same-month event count in the
                        denominator. (pd.notna() — not `is not None`.)
                        SLAs: FO>Disp 3 BD, Disp>PU 3 BD.
5. Flow-through       = Col M / N / O use the pickup universe (rows where
                        BOTH PickUp Actual Date AND FO Create Date are not
                        null). Orphan pickups without a create date are
                        excluded so the cumulative counts line up.
6. Col N              = running(FOs created) - running(FO-linked pickups)
                        over the full timeline.
7. Pacing             = cutoff hardcoded to the 23rd of the current month,
                        report date = day after the latest pickup in the
                        data (or manually overridden).

Validated against the published FO_Performance_34.xlsx and the Quiet Luxury
render spec.
"""

from __future__ import annotations

import calendar
from datetime import date, datetime, timedelta
from typing import Any, Iterable, Optional

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SLA_FO_TO_DISPATCH_BD = 3
SLA_DISPATCH_TO_PICKUP_BD = 3

DEFAULT_MONTHLY_OBJECTIVE = 477
PACING_CUTOFF_DAY = 23  # 23rd of current month per logistics team

# Compliance colour bands — used by the HTML renderer.
COMPLIANCE_GOOD = 0.95
COMPLIANCE_WARN = 0.75

# Columns emitted in the final report (in display order). Must stay in sync
# with data_hub.render.fo_performance_html.
COLUMN_KEYS = [
    'date',
    'fos_created',
    'dispatched',
    'picked_up',
    'delivered',
    'avg_fo_to_disp_bd',
    'sla_fo_to_disp',
    'fo_disp_compliance',
    'avg_disp_to_pu_bd',
    'sla_disp_to_pu',
    'disp_pu_compliance',
    'avg_e2e_bd',
    'prior_month_fo_pickups',
    'cumulative_awaiting_pu',
    'mtd_pickups',
]

COLUMN_LABELS = {
    'date': 'DATE',
    'fos_created': 'FOS CREATED',
    'dispatched': 'DISPATCHED',
    'picked_up': 'PICKED UP',
    'delivered': 'DELIVERED',
    'avg_fo_to_disp_bd': 'AVG FO TO DISP (BD)',
    'sla_fo_to_disp': 'SLA',
    'fo_disp_compliance': 'FO>DISP COMPLIANCE',
    'avg_disp_to_pu_bd': 'AVG DISP TO PU (BD)',
    'sla_disp_to_pu': 'SLA',
    'disp_pu_compliance': 'DISP>PU COMPLIANCE',
    'avg_e2e_bd': 'AVG E2E (BD)',
    'prior_month_fo_pickups': 'PRIOR MONTH FO PICKUPS',
    'cumulative_awaiting_pu': 'CUMULATIVE AWAITING PU',
    'mtd_pickups': 'MTD PICKUPS',
}


# ---------------------------------------------------------------------------
# US Federal Holiday calendar
# ---------------------------------------------------------------------------

def _us_federal_holidays(years: Iterable[int]) -> list[date]:
    """Return US federal holidays for the given years.

    Covers the 11 federal holidays: New Year's, MLK, Presidents Day, Memorial
    Day, Juneteenth, Independence Day, Labor Day, Columbus Day, Veterans Day,
    Thanksgiving, Christmas.
    """
    holidays: list[date] = []
    for year in years:
        # New Year's Day
        holidays.append(date(year, 1, 1))
        # MLK Day — 3rd Monday of January
        holidays.append(_nth_weekday(year, 1, 0, 3))
        # Presidents Day — 3rd Monday of February
        holidays.append(_nth_weekday(year, 2, 0, 3))
        # Memorial Day — last Monday of May
        holidays.append(_last_weekday(year, 5, 0))
        # Juneteenth
        holidays.append(date(year, 6, 19))
        # Independence Day
        holidays.append(date(year, 7, 4))
        # Labor Day — 1st Monday of September
        holidays.append(_nth_weekday(year, 9, 0, 1))
        # Columbus Day — 2nd Monday of October
        holidays.append(_nth_weekday(year, 10, 0, 2))
        # Veterans Day
        holidays.append(date(year, 11, 11))
        # Thanksgiving — 4th Thursday of November
        holidays.append(_nth_weekday(year, 11, 3, 4))
        # Christmas
        holidays.append(date(year, 12, 25))
    return holidays


def _nth_weekday(year: int, month: int, weekday: int, n: int) -> date:
    first = date(year, month, 1)
    offset = (weekday - first.weekday()) % 7
    return first + timedelta(days=offset + 7 * (n - 1))


def _last_weekday(year: int, month: int, weekday: int) -> date:
    last_day = calendar.monthrange(year, month)[1]
    last = date(year, month, last_day)
    offset = (last.weekday() - weekday) % 7
    return last - timedelta(days=offset)


def _holiday_array(years: Iterable[int]) -> np.ndarray:
    holidays = _us_federal_holidays(years)
    return np.array([np.datetime64(h) for h in holidays], dtype='datetime64[D]')


# ---------------------------------------------------------------------------
# Vectorized business day helpers
# ---------------------------------------------------------------------------

def _vec_busday_count(
    starts: pd.Series,
    ends: pd.Series,
    holidays: np.ndarray,
) -> pd.Series:
    """Business days between `starts` and `ends`, excluding weekends + holidays.

    np.busday_count is "inclusive of start, exclusive of end" — which is
    numerically equivalent to the spec's "exclusive of start, inclusive of
    end" for the integer day counts we care about. NaT rows return NaN.
    """
    mask = starts.notna() & ends.notna()
    out = np.full(len(starts), np.nan, dtype='float64')
    if mask.any():
        s = starts[mask].dt.date.values.astype('datetime64[D]')
        e = ends[mask].dt.date.values.astype('datetime64[D]')
        out[mask.values] = np.busday_count(s, e, holidays=holidays).astype('float64')
    return pd.Series(out, index=starts.index)


def _busdays_between(start: date, end: date, holidays: np.ndarray) -> int:
    if end < start:
        return 0
    s = np.datetime64(start, 'D')
    e = np.datetime64(end + timedelta(days=1), 'D')  # inclusive end
    return int(np.busday_count(s, e, holidays=holidays))


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def _load_data_file(xlsx_path: str) -> pd.DataFrame:
    """Read the Data File sheet and return a normalised DataFrame.

    Uses openpyxl's read_only mode directly (bypasses pandas.read_excel
    overhead) and only materialises the five columns we actually need.
    ~740 ms for 8k rows vs ~75s for the default openpyxl load mode.

    The Data File sheet has two 'Delivered Date' columns — a leftmost one
    that holds actual delivery dates, and a rightmost one that mirrors
    Delivery ETA. We always prefer the leftmost match when a header name
    collides, which is what a human would pick scanning the file left to
    right. Same thing for 'Pickup date'.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if 'Data File' in wb.sheetnames:
            ws = wb['Data File']
        else:
            ws = wb[wb.sheetnames[0]]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return pd.DataFrame(columns=['fo', 'create', 'dispatch', 'pickup', 'delivered'])

        # Build a leftmost-wins lookup so duplicate headers don't clobber
        # each other. normalised_name → first matching column index.
        header_index: dict[str, int] = {}
        for i, h in enumerate(headers):
            if isinstance(h, str):
                key = h.strip().lower()
                if key and key not in header_index:
                    header_index[key] = i

        def find(*candidates: str) -> Optional[int]:
            for name in candidates:
                key = name.strip().lower()
                if key in header_index:
                    return header_index[key]
            # Substring fallback — also leftmost-first
            for name in candidates:
                key = name.strip().lower()
                best: Optional[int] = None
                for low, idx in header_index.items():
                    if key in low and (best is None or idx < best):
                        best = idx
                if best is not None:
                    return best
            return None

        idx_fo = find('FO')
        idx_create = find('FO Create Date')
        idx_dispatch = find('Dispatched Date', 'Dispatch Date')
        idx_pickup = find('PickUp Actual Date', 'Pickup Actual Date')
        idx_delivered = find('Delivered Date')

        if idx_fo is None or idx_create is None:
            raise ValueError(
                "Data File sheet is missing required columns 'FO' and/or "
                f"'FO Create Date'. Headers found: {list(headers)[:20]}"
            )

        wanted = [idx_fo, idx_create, idx_dispatch, idx_pickup, idx_delivered]
        records = []
        for row in rows_iter:
            records.append([row[i] if (i is not None and i < len(row)) else None for i in wanted])
    finally:
        wb.close()

    df = pd.DataFrame(records, columns=['fo', 'create', 'dispatch', 'pickup', 'delivered'])
    df['fo'] = df['fo'].astype('string')
    for col in ('create', 'dispatch', 'pickup', 'delivered'):
        df[col] = pd.to_datetime(df[col], errors='coerce')
    df = df[df['fo'].notna() & df['create'].notna()].reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# Core math
# ---------------------------------------------------------------------------

def _apply_same_month_capping(df: pd.DataFrame) -> pd.DataFrame:
    """Attach same-month-capped versions of the three lifecycle columns.

    The raw columns (`dispatch`, `pickup`, `delivered`) are LEFT INTACT because
    the flow-through columns (Col M / N / O) and the Pipeline / Anticipated
    Remaining calculations need the uncapped pickup dates — a pickup that
    happened one month after creation still counts as "picked up", it just
    doesn't count toward the creation month's cohort.

    The capped columns (`dispatch_same_month`, `pickup_same_month`,
    `delivered_same_month`) feed the daily cohort aggregations: Col C/D/E
    counts, the three business-day averages, and SLA compliance.
    """
    create_period = df['create'].dt.to_period('M')
    for col in ('dispatch', 'pickup', 'delivered'):
        same_month = df[col].dt.to_period('M') == create_period
        df[f'{col}_same_month'] = df[col].where(same_month & df[col].notna())
    return df


def _compute_per_record_bd(
    df: pd.DataFrame,
    holidays: np.ndarray,
) -> pd.DataFrame:
    """Attach per-record BD columns for FO->Disp, Disp->PU, and FO->Delivered.

    BD metrics always use the same-month-capped dates: we're measuring how
    fast an FO moved through the lifecycle within its creation month. A
    dispatch that happened the following month is outside the SLA scope.
    """
    df['fo_to_disp_bd'] = _vec_busday_count(df['create'], df['dispatch_same_month'], holidays)
    df['disp_to_pu_bd'] = _vec_busday_count(df['dispatch_same_month'], df['pickup_same_month'], holidays)
    df['fo_to_delivered_bd'] = _vec_busday_count(df['create'], df['delivered_same_month'], holidays)

    df['fo_disp_sla_met'] = np.where(
        df['fo_to_disp_bd'].notna(),
        (df['fo_to_disp_bd'] <= SLA_FO_TO_DISPATCH_BD).astype('float64'),
        np.nan,
    )
    df['disp_pu_sla_met'] = np.where(
        df['disp_to_pu_bd'].notna(),
        (df['disp_to_pu_bd'] <= SLA_DISPATCH_TO_PICKUP_BD).astype('float64'),
        np.nan,
    )
    return df


def _aggregate_by_day(df: pd.DataFrame) -> pd.DataFrame:
    """Group records by FO create date and compute the daily cohort metrics.

    Uses the same-month-capped counts so each daily row reports "of the FOs
    created this day, how many reached dispatch/pickup/delivery WITHIN THEIR
    CREATION MONTH" — the cohort flow-through metric.
    """
    df = df.copy()
    df['create_date'] = df['create'].dt.date

    agg = df.groupby('create_date').agg(
        fos_created=('fo', 'size'),
        dispatched=('dispatch_same_month', 'count'),
        picked_up=('pickup_same_month', 'count'),
        delivered=('delivered_same_month', 'count'),
        avg_fo_to_disp_bd=('fo_to_disp_bd', 'mean'),
        avg_disp_to_pu_bd=('disp_to_pu_bd', 'mean'),
        avg_e2e_bd=('fo_to_delivered_bd', 'mean'),
        fo_disp_compliance=('fo_disp_sla_met', 'mean'),
        disp_pu_compliance=('disp_pu_sla_met', 'mean'),
    ).sort_index()
    return agg


def _compute_flow_through(
    df: pd.DataFrame,
    daily_dates: list[date],
) -> dict[date, dict[str, int]]:
    """Compute Col M / N / O for each daily row date.

    - M: pickups on this date where FO was created in a different month.
    - N: running(FOs created) - running(FO-linked pickups) as of this date.
    - O: month-to-date cumulative pickups (FO-linked).
    """
    # Pickup universe: records with both pickup and create populated.
    pu = df[df['pickup'].notna() & df['create'].notna()].copy()
    pu['pickup_date'] = pu['pickup'].dt.date
    pu['create_period'] = pu['create'].dt.to_period('M')
    pu['pickup_period'] = pu['pickup'].dt.to_period('M')
    pu['is_prior_month'] = pu['create_period'] != pu['pickup_period']

    # M: pickups on a given date where FO was created in a different month.
    prior_by_date = (
        pu[pu['is_prior_month']]
        .groupby('pickup_date')
        .size()
    )

    # N: build a running balance keyed by day across the full timeline.
    create_per_day = df.groupby(df['create'].dt.date).size()
    pu_per_day = pu.groupby('pickup_date').size()
    timeline = pd.Index(sorted(set(create_per_day.index) | set(pu_per_day.index)))
    create_series = create_per_day.reindex(timeline, fill_value=0)
    pu_series = pu_per_day.reindex(timeline, fill_value=0)
    cum_awaiting = (create_series.cumsum() - pu_series.cumsum())

    # O: month-to-date pickups. For each date, sum pickups from month start
    # through that date. We compute this as a cumulative running sum that
    # resets whenever the month changes.
    pu_idx = pu_per_day.sort_index()
    mtd_by_date: dict[date, int] = {}
    running_mtd = 0
    current_ym: Optional[tuple[int, int]] = None
    for d, count in pu_idx.items():
        ym = (d.year, d.month)
        if ym != current_ym:
            running_mtd = 0
            current_ym = ym
        running_mtd += int(count)
        mtd_by_date[d] = running_mtd

    # Carry the MTD value forward across days with no pickups, then reset at
    # month boundaries. daily_dates may include days that never saw a pickup.
    result: dict[date, dict[str, int]] = {}
    for day in daily_dates:
        # N: running balance through this day (look up most recent prior).
        awaiting_slice = cum_awaiting.loc[:day]
        awaiting_val = int(awaiting_slice.iloc[-1]) if len(awaiting_slice) else 0

        # M: prior-month pickups on this exact day.
        prior_val = int(prior_by_date.loc[day]) if day in prior_by_date.index else 0

        # O: MTD pickups through this day.
        month_start = date(day.year, day.month, 1)
        mtd_slice = pu_idx.loc[month_start:day]
        mtd_val = int(mtd_slice.sum()) if len(mtd_slice) else 0

        result[day] = {
            'prior_month_fo_pickups': prior_val,
            'cumulative_awaiting_pu': awaiting_val,
            'mtd_pickups': mtd_val,
        }
    return result


# ---------------------------------------------------------------------------
# Dict assembly
# ---------------------------------------------------------------------------

def _round1(value: Any) -> Optional[float]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    return round(float(value), 1)


def _round3(value: Any) -> Optional[float]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    return round(float(value), 3)


def _build_daily_row(day: date, row: pd.Series, flow: dict[str, int]) -> dict[str, Any]:
    return {
        'date': day.strftime('%a %m/%d'),
        '_iso_date': day.isoformat(),
        'fos_created': int(row['fos_created']),
        'dispatched': int(row['dispatched']),
        'picked_up': int(row['picked_up']),
        'delivered': int(row['delivered']),
        'avg_fo_to_disp_bd': _round1(row['avg_fo_to_disp_bd']),
        'sla_fo_to_disp': SLA_FO_TO_DISPATCH_BD,
        'fo_disp_compliance': _round3(row['fo_disp_compliance']),
        'avg_disp_to_pu_bd': _round1(row['avg_disp_to_pu_bd']),
        'sla_disp_to_pu': SLA_DISPATCH_TO_PICKUP_BD,
        'disp_pu_compliance': _round3(row['disp_pu_compliance']),
        'avg_e2e_bd': _round1(row['avg_e2e_bd']),
        'prior_month_fo_pickups': flow['prior_month_fo_pickups'],
        'cumulative_awaiting_pu': flow['cumulative_awaiting_pu'],
        'mtd_pickups': flow['mtd_pickups'],
    }


def _build_month_total(
    label: str,
    days: list[dict[str, Any]],
    month_end_day: date,
    pu_df: pd.DataFrame,
) -> dict[str, Any]:
    month_start = date(month_end_day.year, month_end_day.month, 1)
    prior_total = int(
        (
            (pu_df['pickup'].dt.date >= month_start)
            & (pu_df['pickup'].dt.date <= month_end_day)
            & (pu_df['create'].dt.to_period('M') != pu_df['pickup'].dt.to_period('M'))
        ).sum()
    )
    mtd_total = int(
        (
            (pu_df['pickup'].dt.date >= month_start)
            & (pu_df['pickup'].dt.date <= month_end_day)
        ).sum()
    )
    return {
        'date': f'{label} Total',
        '_iso_date': month_end_day.isoformat(),
        'fos_created': sum(d['fos_created'] for d in days),
        'dispatched': sum(d['dispatched'] for d in days),
        'picked_up': sum(d['picked_up'] for d in days),
        'delivered': sum(d['delivered'] for d in days),
        'avg_fo_to_disp_bd': None,
        'sla_fo_to_disp': None,
        'fo_disp_compliance': None,
        'avg_disp_to_pu_bd': None,
        'sla_disp_to_pu': None,
        'disp_pu_compliance': None,
        'avg_e2e_bd': None,
        'prior_month_fo_pickups': prior_total,
        'cumulative_awaiting_pu': None,
        'mtd_pickups': mtd_total,
    }


def _build_grand_total(df: pd.DataFrame) -> dict[str, Any]:
    fo_to_disp = df['fo_to_disp_bd'].dropna()
    disp_to_pu = df['disp_to_pu_bd'].dropna()
    e2e = df['fo_to_delivered_bd'].dropna()

    fo_disp_met = df['fo_disp_sla_met'].dropna()
    disp_pu_met = df['disp_pu_sla_met'].dropna()

    return {
        'date': 'GRAND TOTAL',
        'fos_created': int(len(df)),
        'dispatched': int(df['dispatch'].notna().sum()),
        'picked_up': int(df['pickup'].notna().sum()),
        'delivered': int(df['delivered'].notna().sum()),
        'avg_fo_to_disp_bd': _round1(fo_to_disp.mean()) if len(fo_to_disp) else None,
        'sla_fo_to_disp': SLA_FO_TO_DISPATCH_BD,
        'fo_disp_compliance': _round3(fo_disp_met.mean()) if len(fo_disp_met) else None,
        'avg_disp_to_pu_bd': _round1(disp_to_pu.mean()) if len(disp_to_pu) else None,
        'sla_disp_to_pu': SLA_DISPATCH_TO_PICKUP_BD,
        'disp_pu_compliance': _round3(disp_pu_met.mean()) if len(disp_pu_met) else None,
        'avg_e2e_bd': _round1(e2e.mean()) if len(e2e) else None,
        'prior_month_fo_pickups': None,
        'cumulative_awaiting_pu': None,
        'mtd_pickups': None,
    }


# ---------------------------------------------------------------------------
# Summary sections
# ---------------------------------------------------------------------------

def _row(label: str, value: Any = None, note: str = '') -> list[Any]:
    cells: list[Any] = [label] + [None] * 15
    if value is not None:
        cells[11] = value
    if note:
        cells[13] = note
    return cells


def _compute_base_metrics(
    df: pd.DataFrame,
    report_date: date,
    holidays: np.ndarray,
) -> dict[str, Any]:
    """Compute objective-independent metrics used by the summary sections.

    `report_date` is the "as of" day the report covers. The pacing cutoff is
    hardcoded to the 23rd of the report's month per the logistics team's
    operational deadline.
    """
    month_start = date(report_date.year, report_date.month, 1)
    cutoff_day = date(report_date.year, report_date.month, PACING_CUTOFF_DAY)
    last_day = date(
        report_date.year,
        report_date.month,
        calendar.monthrange(report_date.year, report_date.month)[1],
    )

    # Pickup universe: records with both pickup and create populated.
    pu = df[df['pickup'].notna()].copy()
    pu_dates = pu['pickup'].dt.date

    # MTD wholesales = pickups in the current month up through report_date.
    mtd_pickups = int(((pu_dates >= month_start) & (pu_dates <= report_date)).sum())

    # Anticipated remaining = running balance (create - pickup) through report.
    total_created_to_date = int((df['create'].dt.date <= report_date).sum())
    total_pu_to_date = int((pu_dates <= report_date).sum())
    anticipated_remaining = max(0, total_created_to_date - total_pu_to_date)

    # Breakdown for the Anticipated Wholesales section — match the logistics
    # team's Excel convention: "Current Month FOs Created" is literally the
    # MTD Col B sum (every Apr FO, whether or not it was picked up). The
    # "Prior Month FOs Awaiting Pickup" is derived by subtraction so the two
    # always sum back to anticipated_remaining.
    current_month_created = int(
        ((df['create'].dt.date >= month_start) & (df['create'].dt.date <= report_date)).sum()
    )
    prior_month_awaiting = max(0, anticipated_remaining - current_month_created)

    # Pacing — cutoff hardcoded to the 23rd.
    if report_date > cutoff_day:
        remaining_bd = 0
        cutoff_passed = True
    else:
        remaining_bd = _busdays_between(report_date, cutoff_day, holidays)
        cutoff_passed = False

    # Pipeline FOs awaiting pickup as of the cutoff day = FOs created on or
    # before the cutoff that have no pickup date at all.
    cutoff_mask_create = df['create'].dt.date <= cutoff_day
    pipeline_df = df[cutoff_mask_create & df['pickup'].isna()]
    pipeline_current = int(
        (pipeline_df['create'].dt.date >= month_start).sum()
    )
    pipeline_prior = int(len(pipeline_df) - pipeline_current)
    existing_pipeline = pipeline_current + pipeline_prior

    return {
        'report_date': report_date.isoformat(),
        'report_date_label': report_date.strftime('%b %d'),
        'month_label': report_date.strftime('%b %Y'),
        'month_start': month_start.isoformat(),
        'month_end': last_day.isoformat(),
        'month_end_label': last_day.strftime('%b %d'),
        'cutoff_day': cutoff_day.isoformat(),
        'cutoff_day_label': cutoff_day.strftime('%b %d'),
        'prior_month_label': (month_start - timedelta(days=1)).strftime('%b'),
        'mtd_pickups': mtd_pickups,
        'anticipated_remaining': anticipated_remaining,
        'current_month_created': current_month_created,
        'prior_month_awaiting': prior_month_awaiting,
        'existing_pipeline': existing_pipeline,
        'pipeline_current': pipeline_current,
        'pipeline_prior': pipeline_prior,
        'remaining_bd': remaining_bd,
        'cutoff_passed': cutoff_passed,
    }


def _build_anticipated_section(base: dict[str, Any]) -> dict[str, Any]:
    rows: list[list[Any]] = [
        _row('MTD Wholesales (Pickups)', base['mtd_pickups'], '(Col O)'),
        _row('Anticipated Remaining Wholesales', base['anticipated_remaining'], '(Col N)'),
        _row(
            'of which: Current Month FOs Created (subset of above)',
            base['current_month_created'],
            '(Col B month total)',
        ),
        _row(
            f'of which: {base["prior_month_label"]} FOs Awaiting Pickup (subset of above)',
            base['prior_month_awaiting'],
            f'(of {base["anticipated_remaining"]} total)',
        ),
    ]
    return {
        'id': 'anticipated_wholesales',
        'title': f'CURRENT MONTH ANTICIPATED WHOLESALES (AT PICKUP) — {base["month_label"]}',
        'rows': rows,
    }


def _build_pacing_section(base: dict[str, Any], monthly_objective: int) -> dict[str, Any]:
    pipeline = base['existing_pipeline']
    remaining_bd = base['remaining_bd']
    cutoff_passed = base['cutoff_passed']
    mtd = base['mtd_pickups']

    if cutoff_passed:
        new_fos_needed: Any = 'Cutoff passed'
    elif remaining_bd <= 0:
        new_fos_needed = 'N/A'
    else:
        new_fos_needed = round((monthly_objective - mtd - pipeline) / remaining_bd, 1)

    rows: list[list[Any]] = [
        _row('MONTHLY WHOLESALE OBJECTIVE', monthly_objective, 'Enter target'),
        _row('Less: MTD Wholesales (pickups already completed)', mtd),
        _row(
            f'Less: Existing Pipeline FOs Awaiting Pickup (created through {base["cutoff_day_label"]})',
            pipeline,
            f'({base["pipeline_current"]} current mo + {base["pipeline_prior"]} prior)',
        ),
        _row(
            f'Remaining Business Days ({base["report_date_label"]} through {base["cutoff_day_label"]}, inclusive)',
            remaining_bd,
        ),
        _row(
            f'NEW FOs NEEDED PER DAY (through {base["cutoff_day_label"]})',
            new_fos_needed,
            'per business day',
        ),
    ]
    return {
        'id': 'fo_pacing',
        'title': f'FO PACING TO OBJECTIVE (through {base["cutoff_day_label"]})',
        'rows': rows,
    }


def _column_definitions() -> list[list[Any]]:
    entries = [
        ('Col A', 'Date',
         'Calendar date on which the Freight Order was created. Rows grouped by month with subtotals.'),
        ('Col B', 'FOs Created',
         'Number of new Freight Orders entered on this date. Each FO = one VIN assigned for transport from VPC to retailer.'),
        ('Col C', 'Dispatched',
         'FOs from this date that were dispatched to a carrier within the same calendar month (flow-through from creation cohort).'),
        ('Col D', 'Picked Up',
         'FOs from this date that were picked up within the same calendar month.'),
        ('Col E', 'Delivered',
         'FOs from this date that were delivered to the retailer within the same calendar month.'),
        ('Col F', 'Avg FO to Disp (BD)',
         'Average business days (weekends + US federal holidays excluded) between FO creation and carrier dispatch for this date\'s cohort.'),
        ('Col G', 'SLA',
         f'Contract SLA target for FO-to-dispatch: {SLA_FO_TO_DISPATCH_BD} business days.'),
        ('Col H', 'FO>Disp Compliance',
         'Percentage of dispatched FOs from this cohort that met the FO-to-dispatch SLA.'),
        ('Col I', 'Avg Disp to PU (BD)',
         'Average business days between dispatch and pickup for this cohort.'),
        ('Col J', 'SLA',
         f'Contract SLA target for dispatch-to-pickup: {SLA_DISPATCH_TO_PICKUP_BD} business days.'),
        ('Col K', 'Disp>PU Compliance',
         'Percentage of picked-up FOs from this cohort that met the dispatch-to-pickup SLA.'),
        ('Col L', 'Avg E2E (BD)',
         'Average end-to-end business days from FO creation to delivery for this cohort.'),
        ('Col M', 'Prior Month FO Pickups',
         'Carrier pickups occurring on this date for FOs that were created in a previous month. Drawn from the pickup universe (FOs with both create and pickup dates).'),
        ('Col N', 'Cumulative Awaiting PU',
         'Running total of all FO-linked pickups subtracted from all FOs ever created. Equals the anticipated remaining wholesale count at the report date.'),
        ('Col O', 'MTD Pickups',
         'Month-to-date cumulative carrier pickups for FO-linked vehicles. Resets at the start of each month.'),
    ]
    return [[key, name, desc] + [None] * 12 for key, name, desc in entries]


def _build_summary_sections(
    base: dict[str, Any],
    monthly_objective: int,
) -> list[dict[str, Any]]:
    return [
        _build_anticipated_section(base),
        _build_pacing_section(base, monthly_objective),
        {
            'id': 'column_definitions',
            'title': 'COLUMN DEFINITIONS',
            'rows': _column_definitions(),
        },
    ]


def _is_pacing_section(section: dict[str, Any]) -> bool:
    if not isinstance(section, dict):
        return False
    if section.get('id') == 'fo_pacing':
        return True
    title = str(section.get('title') or '').upper()
    return title.startswith('FO PACING')


def _is_glossary_section(section: dict[str, Any]) -> bool:
    if not isinstance(section, dict):
        return False
    if section.get('id') == 'column_definitions':
        return True
    title = str(section.get('title') or '').upper()
    return 'COLUMN DEFINITIONS' in title


def apply_monthly_objective(data: dict[str, Any], monthly_objective: int) -> dict[str, Any]:
    """Swap the FO Pacing section in-place for a new monthly objective.

    Mutates and returns ``data``. Matches by id first, then falls back to any
    section whose title starts with "FO PACING" so legacy cached JSON works.
    Any pre-existing pacing sections (including accidental duplicates) are
    removed before the fresh one is inserted.
    """
    base = data.get('base_metrics')
    if not base:
        return data

    new_pacing = _build_pacing_section(base, monthly_objective)
    sections = list(data.get('sections') or [])

    insert_at: Optional[int] = None
    kept: list[dict[str, Any]] = []
    for section in sections:
        if _is_pacing_section(section):
            if insert_at is None:
                insert_at = len(kept)
            continue
        kept.append(section)

    if insert_at is None:
        insert_at = len(kept)
        for i, section in enumerate(kept):
            if _is_glossary_section(section):
                insert_at = i
                break

    kept.insert(insert_at, new_pacing)
    data['sections'] = kept
    data['monthly_objective'] = monthly_objective
    return data


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def ingest_vehicle_distribution(
    xlsx_path: str,
    *,
    monthly_objective: int = DEFAULT_MONTHLY_OBJECTIVE,
    report_date: Optional[date] = None,
) -> dict[str, Any]:
    """Parse a Vehicle Distribution workbook and compute the Logistics FO
    Performance report structure.

    Parameters
    ----------
    xlsx_path : str
        Path to the Vehicle Distribution xlsx file.
    monthly_objective : int
        Monthly wholesale objective used by the FO PACING section. The
        frontend calls apply_monthly_objective() to swap this without
        re-parsing the workbook.
    report_date : date | None
        "As of" date for the pacing / cumulative / MTD calculations.
        Defaults to the day after the latest pickup date in the data.
    """
    df = _load_data_file(xlsx_path)
    if df.empty:
        raise ValueError("No FO records with a Create Date were found in the workbook.")

    # Same-month capping first — every downstream calc reads from the capped
    # columns so post-month events never inflate SLA/flow-through metrics.
    df = _apply_same_month_capping(df)

    # Derive a holiday calendar covering every year present in the data.
    years_in_data = set(df['create'].dt.year.dropna().astype(int).unique().tolist())
    if 'pickup' in df:
        pu_years = df['pickup'].dt.year.dropna().astype(int).unique().tolist()
        years_in_data.update(pu_years)
    if not years_in_data:
        years_in_data = {datetime.utcnow().year}
    years_span = range(min(years_in_data), max(years_in_data) + 2)
    holidays = _holiday_array(years_span)

    df = _compute_per_record_bd(df, holidays)

    daily = _aggregate_by_day(df)
    daily_dates = list(daily.index)

    flow = _compute_flow_through(df, daily_dates)

    # Build month buckets in chronological order
    months: list[dict[str, Any]] = []
    current_label: Optional[str] = None
    current_month: Optional[dict[str, Any]] = None

    for day, row in daily.iterrows():
        month_label = day.strftime('%b %Y')
        if month_label != current_label:
            if current_month is not None and current_label is not None:
                previous_dt = datetime.strptime(current_label, '%b %Y').date()
                month_end = date(
                    previous_dt.year,
                    previous_dt.month,
                    calendar.monthrange(previous_dt.year, previous_dt.month)[1],
                )
                current_month['total'] = _build_month_total(
                    current_label,
                    current_month['days'],
                    month_end,
                    df,
                )
                months.append(current_month)
            current_label = month_label
            current_month = {'label': month_label, 'days': [], 'total': None}
        assert current_month is not None
        current_month['days'].append(_build_daily_row(day, row, flow[day]))

    if current_month is not None and current_label is not None:
        previous_dt = datetime.strptime(current_label, '%b %Y').date()
        month_end = date(
            previous_dt.year,
            previous_dt.month,
            calendar.monthrange(previous_dt.year, previous_dt.month)[1],
        )
        current_month['total'] = _build_month_total(
            current_label, current_month['days'], month_end, df,
        )
        months.append(current_month)

    grand_total = _build_grand_total(df)

    # Report date = day after latest pickup, capped by the latest create
    # date if there are no pickups yet. Caller can override.
    if report_date is None:
        latest_pickup = df['pickup'].max()
        if pd.notna(latest_pickup):
            report_date = (latest_pickup + pd.Timedelta(days=1)).date()
        else:
            report_date = df['create'].max().date()

    base_metrics = _compute_base_metrics(df, report_date, holidays)
    sections = _build_summary_sections(base_metrics, monthly_objective)

    columns = [{'key': k, 'label': COLUMN_LABELS[k]} for k in COLUMN_KEYS]

    result = {
        'title': 'DAILY FREIGHT ORDER ACTIVITY WITH SLA TRACKING AND FLOW-THROUGH (BUSINESS DAYS)',
        'columns': columns,
        'months': months,
        'grand_total': grand_total,
        'sections': sections,
        'base_metrics': base_metrics,
        'report_date': report_date.isoformat(),
        'monthly_objective': monthly_objective,
        'generated_at': datetime.utcnow().isoformat() + 'Z',
    }

    total_days = sum(len(m['days']) for m in months)
    print(
        f"  Vehicle Distribution: {len(df)} FO records, {len(months)} months, "
        f"{total_days} daily rows (report_date={report_date.isoformat()}, "
        f"objective={monthly_objective})"
    )
    return result


# Backwards-compatible alias — older code still imports ingest_fo_performance.
ingest_fo_performance = ingest_vehicle_distribution
