"""FO Performance (Logistics Freight Order) ingest.

Parses the 'Daily Freight Order Activity with SLA Tracking and Flow-Through'
xlsx report into a structured JSON-serialisable dict. The input workbook has a
single sheet 'FO Performance' laid out as:

  row 0: banner title (merged)
  row 1: column headers (DATE, FOS CREATED, DISPATCHED, PICKED UP, DELIVERED,
         AVG FO TO DISP, SLA, FO>DISP COMPLIANCE, AVG DISP TO PU, SLA,
         DISP>PU COMPLIANCE, AVG E2E, PRIOR MONTH FO PICKUPS,
         CUMULATIVE AWAITING PU, MTD PICKUPS)
  rows 2..N: daily rows grouped by month, with a bold "<Mon YYYY> Total" row
             at the end of each month. Trailing months may have blank spacer
             rows between them.
  ...          : GRAND TOTAL row
  ...          : "CURRENT MONTH ANTICIPATED WHOLESALES (AT PICKUP)" summary
                 block with 4-5 labelled lines
  ...          : "FO PACING TO OBJECTIVE (through <date>)" summary block
  ...          : "COLUMN DEFINITIONS" reference block (glossary)

The parser walks the sheet once, classifies each row as one of:
  - banner  : top title row (extracted into `title`)
  - header  : column header row (extracted into `columns`)
  - daily   : per-day data row (added to `months[...]['days']`)
  - total   : month total row (stored as `months[...]['total']`)
  - grand   : grand total row (stored as `grand_total`)
  - section : summary section header ("CURRENT MONTH ...", "FO PACING ...",
              "COLUMN DEFINITIONS") — creates a new block
  - section_row : body row inside a summary block
  - blank   : skipped

The resulting dict is JSON-friendly (all cell values coerced to int / float /
string) so it can be stored in the CachedFile table and rendered to HTML by
`data_hub.render.fo_performance_html` without re-reading the xlsx.
"""

from __future__ import annotations

import re
from datetime import datetime, date
from typing import Any

import openpyxl


# Internal column keys matching the Excel layout. Order matters — this is
# also the order the HTML renderer uses when writing table cells.
COLUMN_KEYS = [
    'date',
    'fos_created',
    'dispatched',
    'picked_up',
    'delivered',
    'avg_fo_to_disp_bd',
    'sla_fo_to_disp',
    'fo_disp_compliance',
    'avg_disp_to_pu_bd',
    'sla_disp_to_pu',
    'disp_pu_compliance',
    'avg_e2e_bd',
    'prior_month_fo_pickups',
    'cumulative_awaiting_pu',
    'mtd_pickups',
]


_TOTAL_RE = re.compile(r'^\s*([A-Za-z]{3,9})\s+(\d{4})\s+total\s*$', re.IGNORECASE)
_GRAND_RE = re.compile(r'^\s*grand\s+total\s*$', re.IGNORECASE)
_SECTION_HEADERS = {
    'CURRENT MONTH ANTICIPATED WHOLESALES': 'anticipated_wholesales',
    'FO PACING TO OBJECTIVE': 'fo_pacing',
    'COLUMN DEFINITIONS': 'column_definitions',
}


def _cell_to_json(value: Any) -> Any:
    """Coerce a cell value into something JSON-serialisable."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    # Normalise embedded newlines that Excel keeps inside header cells.
    s = s.replace('\r', '').replace('\n', ' ').strip()
    return s


def _is_number(value: Any) -> bool:
    if isinstance(value, bool):  # bool is subclass of int — exclude explicitly
        return False
    return isinstance(value, (int, float))


def _row_is_blank(row: list[Any]) -> bool:
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)


def _classify_row(first_cell: Any, row: list[Any]) -> str:
    """Classify a row based on its first cell and remaining values."""
    if _row_is_blank(row):
        return 'blank'
    if first_cell is None:
        return 'blank'
    text = str(first_cell).strip()
    if not text:
        return 'blank'
    upper = text.upper()
    if _GRAND_RE.match(text):
        return 'grand'
    if _TOTAL_RE.match(text):
        return 'total'
    for prefix in _SECTION_HEADERS:
        if upper.startswith(prefix):
            return 'section'
    if upper == 'DATE' or upper.startswith('DATE\n') or upper == 'DATE ':
        return 'header'
    if 'DAILY FREIGHT ORDER ACTIVITY' in upper:
        return 'banner'
    # A "daily" row typically looks like "Mon 03/31" or "Tue 07/15". Any
    # row whose first cell is a date-like string or datetime and whose
    # second cell is numeric is treated as a daily entry.
    if isinstance(first_cell, (datetime, date)):
        return 'daily'
    # Heuristic: first cell has a short weekday prefix and the 2nd cell is
    # numeric → daily row.
    if len(row) > 1 and _is_number(row[1]):
        return 'daily'
    # Otherwise treat as a summary / body row under the current section.
    return 'section_row'


def _build_daily_entry(row: list[Any]) -> dict[str, Any]:
    entry: dict[str, Any] = {}
    for i, key in enumerate(COLUMN_KEYS):
        val = row[i] if i < len(row) else None
        entry[key] = _cell_to_json(val)
    return entry


def _build_total_entry(row: list[Any]) -> dict[str, Any]:
    # Same columns but with total row styling handled downstream.
    return _build_daily_entry(row)


def ingest_fo_performance(xlsx_path: str) -> dict[str, Any]:
    """Parse the FO Performance xlsx and return a structured dict.

    Returns
    -------
    {
        'title':     'DAILY FREIGHT ORDER ACTIVITY ...',
        'columns':   [{'key': 'date', 'label': 'Date'}, ...],
        'months':    [
            {'label': 'Jul 2025', 'days': [<entry>, ...], 'total': <entry>},
            ...
        ],
        'grand_total': <entry> | None,
        'sections': [
            {'title': 'CURRENT MONTH ...', 'rows': [[c0, c1, ...], ...]},
            ...
        ],
        'generated_at': 'ISO-8601',
    }
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'FO Performance' in wb.sheetnames:
        ws = wb['FO Performance']
    else:
        # Fall back to the first sheet if the exact name is different.
        ws = wb[wb.sheetnames[0]]

    title: str | None = None
    header_row: list[Any] | None = None
    months: list[dict[str, Any]] = []
    current_month: dict[str, Any] | None = None
    grand_total: dict[str, Any] | None = None
    sections: list[dict[str, Any]] = []
    current_section: dict[str, Any] | None = None

    for row_tuple in ws.iter_rows(values_only=True):
        row = list(row_tuple)
        first = row[0] if row else None
        kind = _classify_row(first, row)

        if kind == 'banner':
            title = str(first).strip() if first else None
            continue
        if kind == 'header':
            header_row = row
            continue
        if kind == 'blank':
            # Blank row terminates a running month block but does not close
            # a running summary section (summary sections rely on an explicit
            # next-section header or the end of the sheet).
            current_month = None
            continue
        if kind == 'daily':
            if current_section is not None:
                # A daily-looking row appearing inside a summary section
                # should still belong to the section, not start a new month.
                current_section['rows'].append([_cell_to_json(c) for c in row])
                continue
            if current_month is None:
                current_month = {'label': '', 'days': [], 'total': None}
                months.append(current_month)
            current_month['days'].append(_build_daily_entry(row))
            continue
        if kind == 'total':
            m = _TOTAL_RE.match(str(first))
            label = f"{m.group(1).title()} {m.group(2)}" if m else str(first).strip()
            if current_month is None:
                current_month = {'label': label, 'days': [], 'total': None}
                months.append(current_month)
            else:
                current_month['label'] = label
            current_month['total'] = _build_total_entry(row)
            current_month = None  # next daily row starts a new month
            continue
        if kind == 'grand':
            grand_total = _build_total_entry(row)
            current_month = None
            continue
        if kind == 'section':
            label = str(first).strip()
            current_section = {'title': label, 'rows': []}
            sections.append(current_section)
            continue
        if kind == 'section_row':
            if current_section is None:
                # Orphan row before any section header — create a catch-all.
                current_section = {'title': 'Notes', 'rows': []}
                sections.append(current_section)
            current_section['rows'].append([_cell_to_json(c) for c in row])
            continue

    # Build the columns list from the header row if we captured one,
    # otherwise fall back to the canonical labels.
    if header_row is not None:
        labels = [_cell_to_json(c) or '' for c in header_row[:len(COLUMN_KEYS)]]
        columns = [{'key': k, 'label': labels[i] if i < len(labels) and labels[i] else k.replace('_', ' ').title()}
                   for i, k in enumerate(COLUMN_KEYS)]
    else:
        columns = [{'key': k, 'label': k.replace('_', ' ').title()} for k in COLUMN_KEYS]

    # Filter out empty months (can happen when the parser sees a lone
    # spacer row that never got a total).
    months = [m for m in months if m['days'] or m['total']]

    result = {
        'title': title or 'Daily Freight Order Activity',
        'columns': columns,
        'months': months,
        'grand_total': grand_total,
        'sections': sections,
        'generated_at': datetime.utcnow().isoformat() + 'Z',
    }

    total_days = sum(len(m['days']) for m in months)
    print(f"  FO Performance: {len(months)} months, {total_days} daily rows, "
          f"{len(sections)} summary sections")
    return result
