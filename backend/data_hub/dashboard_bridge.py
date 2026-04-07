"""Dashboard Bridge — runs the original dashboard_refresh_all_in_one.py processor
using an assembled xlsx workbook instead of the encrypted Master File.

Strategy:
1. master_assembler.py creates a temp .xlsx from uploaded source DataFrames
2. This bridge opens that .xlsx with openpyxl
3. An adapter wraps openpyxl to mimic pyxlsb's workbook interface
4. The original 3400-line processor runs UNCHANGED against the adapter
5. Result: identical Dashboard HTML as if the Master File were used

This ensures 100% format compatibility because the same battle-tested
code generates the output — no re-implementation of 35 const generators."""

import os
import sys
import io
import types
import tempfile
import importlib
import importlib.util
import openpyxl
from datetime import datetime


# ═══════════════════════════════════════════════════════════════════════
# Adapter: makes openpyxl workbook look like pyxlsb workbook
# ═══════════════════════════════════════════════════════════════════════

class _Cell:
    """Mimics pyxlsb Cell with .v attribute."""
    __slots__ = ('v',)

    def __init__(self, value):
        self.v = value


class _SheetAdapter:
    """Mimics pyxlsb sheet context manager.
    Yields rows of _Cell objects from an openpyxl worksheet."""

    def __init__(self, openpyxl_ws):
        self._ws = openpyxl_ws

    def __enter__(self):
        return self

    def __exit__(self, *args):
        pass

    def rows(self):
        if self._ws is None:
            return
        for row in self._ws.iter_rows():
            # CRITICAL: pyxlsb always returns floats for numeric cells.
            # openpyxl returns ints for whole numbers (46082 instead of 46082.0).
            # The processor uses isinstance(value, float) checks which fail
            # on ints, causing it to skip data. Convert ints to floats here.
            yield [_Cell(float(c.value) if isinstance(c.value, int) and not isinstance(c.value, bool) else c.value) for c in row]


class _EmptySheet:
    """Empty sheet for when a sheet name doesn't exist."""
    def __enter__(self):
        return self

    def __exit__(self, *args):
        pass

    def rows(self):
        return iter([])


class XlsxWorkbookAdapter:
    """Wraps an openpyxl Workbook to present the pyxlsb interface.

    Usage:
        wb = XlsxWorkbookAdapter(openpyxl.load_workbook('file.xlsx'))
        with wb.get_sheet("Export") as sheet:
            for row in sheet.rows():
                vals = [c.v for c in row]
    """

    def __init__(self, openpyxl_wb):
        self._wb = openpyxl_wb

    def get_sheet(self, name):
        if name in self._wb.sheetnames:
            return _SheetAdapter(self._wb[name])
        # Try case-insensitive match
        for sn in self._wb.sheetnames:
            if sn.lower() == name.lower():
                return _SheetAdapter(self._wb[sn])
        print(f"  [Bridge] Sheet '{name}' not found, returning empty sheet")
        return _EmptySheet()

    @property
    def sheets(self):
        return self._wb.sheetnames

    def close(self):
        self._wb.close()


# ═══════════════════════════════════════════════════════════════════════
# Bridge: run original processor against assembled xlsx
# ═══════════════════════════════════════════════════════════════════════

def run_processor_on_xlsx(xlsx_path, template_path, output_path):
    """Run the original dashboard processor against an assembled .xlsx file.

    Instead of decrypting a .xlsb Master File and using pyxlsb,
    we open the .xlsx with openpyxl and wrap it in an adapter that
    presents the same interface. The processor code runs unchanged.

    Args:
        xlsx_path: Path to the assembled .xlsx workbook
        template_path: Path to the Dashboard HTML template
        output_path: Path where the generated Dashboard HTML will be written

    Returns:
        dict with status, output_path, file_size, timestamp
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Assembled workbook not found: {xlsx_path}")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print(f"\n[Bridge] Loading assembled workbook: {xlsx_path}")
    print(f"[Bridge] Template: {template_path}")
    print(f"[Bridge] Output: {output_path}")

    # Open the xlsx with openpyxl and wrap in adapter
    openpyxl_wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    adapter_wb = XlsxWorkbookAdapter(openpyxl_wb)

    print(f"[Bridge] Sheets found: {adapter_wb.sheets}")

    # Load the processor module with monkey-patched imports
    # so it uses our adapter instead of pyxlsb
    processor_path = os.path.join(os.path.dirname(__file__), 'processor_original.py')
    if not os.path.exists(processor_path):
        raise FileNotFoundError(
            f"Processor not found at {processor_path}. "
            "Copy dashboard_refresh_all_in_one.py to data_hub/processor_original.py"
        )

    # Read the processor source and modify it to skip the pyxlsb/msoffcrypto imports
    # and the decrypt step. We'll inject our adapter directly.
    proc_module = _load_processor_module(processor_path)

    # Override decrypt_master to return the input path unchanged
    proc_module.decrypt_master = lambda path, **kw: path

    # Override open_workbook to return our adapter
    proc_module.open_workbook = lambda path: adapter_wb

    # Set sys.argv for main()
    old_argv = sys.argv[:]
    try:
        sys.argv = ["", xlsx_path, template_path, output_path]
        proc_module.main()
    finally:
        sys.argv = old_argv
        openpyxl_wb.close()

    # Verify output
    if not os.path.exists(output_path):
        raise RuntimeError("Processor did not produce output file")

    file_size = os.path.getsize(output_path)
    print(f"\n[Bridge] Dashboard generated: {file_size:,} bytes")

    return {
        'status': 'success',
        'output_path': output_path,
        'file_size': file_size,
        'timestamp': datetime.now().isoformat(),
    }


def _load_processor_module(processor_path):
    """Load the processor as a Python module, handling import substitutions.

    We need to intercept 'from pyxlsb import open_workbook' since:
    1. pyxlsb may not be needed (we use openpyxl adapter)
    2. We want to override open_workbook anyway

    Strategy: Load normally (pyxlsb IS in requirements.txt), then override.
    """
    # Use unique module name to avoid caching issues
    module_name = f'_dashboard_processor_{id(processor_path)}'

    spec = importlib.util.spec_from_file_location(module_name, processor_path)
    module = importlib.util.module_from_spec(spec)

    # We need pyxlsb to be importable so the module loads,
    # but we'll override open_workbook after loading
    try:
        spec.loader.exec_module(module)
    except ImportError as e:
        # If pyxlsb/msoffcrypto not installed, create stubs
        if 'pyxlsb' in str(e):
            _install_fake_pyxlsb()
            spec.loader.exec_module(module)
        elif 'msoffcrypto' in str(e):
            _install_fake_msoffcrypto()
            spec.loader.exec_module(module)
        else:
            raise

    return module


def _install_fake_pyxlsb():
    """Install a fake pyxlsb module if the real one isn't available."""
    fake = types.ModuleType('pyxlsb')
    fake.open_workbook = lambda path: None
    sys.modules['pyxlsb'] = fake


def _install_fake_msoffcrypto():
    """Install a fake msoffcrypto module if the real one isn't available."""
    fake = types.ModuleType('msoffcrypto')

    class FakeOfficeFile:
        def __init__(self, f):
            pass
        def load_key(self, **kw):
            pass
        def decrypt(self, buf):
            pass

    fake.OfficeFile = FakeOfficeFile
    sys.modules['msoffcrypto'] = fake


# ═══════════════════════════════════════════════════════════════════════
# High-level: assemble + process in one call
# ═══════════════════════════════════════════════════════════════════════

def generate_dashboard_from_sources(cache_dir, template_path, output_path):
    """Full pipeline: assemble xlsx from cached sources → run processor → write HTML.

    This is the main entry point called by the orchestrator.

    Args:
        cache_dir: Path to the cache directory with data/*.parquet files
        template_path: Path to the Dashboard HTML template
        output_path: Path where the generated Dashboard HTML will be written

    Returns:
        dict with status, output_path, file_size, vehicle_count, timestamp
    """
    from data_hub.master_assembler import assemble_master_xlsx

    print("\n" + "=" * 60)
    print("DASHBOARD GENERATION FROM SOURCE FILES")
    print("=" * 60)

    # Step 1: Assemble xlsx from cached parquet files
    print("\nStep 1: Assembling workbook from source data...")
    xlsx_path = assemble_master_xlsx(cache_dir, template_path=template_path)

    try:
        # Step 2: Run the original processor against the assembled xlsx
        print("\nStep 2: Running Dashboard processor...")
        result = run_processor_on_xlsx(xlsx_path, template_path, output_path)

        # Count vehicles for reporting
        import pandas as pd
        sap_path = os.path.join(cache_dir, 'data', 'sap_export.parquet')
        vehicle_count = 0
        if os.path.exists(sap_path):
            vehicle_count = len(pd.read_parquet(sap_path))

        result['vehicle_count'] = vehicle_count
        print(f"\n[Bridge] Complete! {vehicle_count} vehicles processed.")
        return result

    finally:
        # Clean up temp file
        try:
            os.unlink(xlsx_path)
        except OSError:
            pass
